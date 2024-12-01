# src/leaf_area_classification.py

#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Leaf Area Classification Processing Script with Adaptive Thresholding in RGB Space

This script processes leaf images to calculate leaf area using adaptive thresholding
in RGB space applied immediately after cropping, before converting to grayscale.

Features:
---------
- Adaptive thresholding in RGB space to exclude near-white non-leaf areas.
- Cropping based on configurable percentages.
- Optional contrast adjustment using CLAHE.
- Morphological operations to clean up the binary image.
- Contour detection and area calculation.
- Mean RGB calculation within detected leaf areas.
- Configurable via 'config.ini' file.
- Supports processing of individual files or entire directories.
- Graceful termination using threading events.
- Annotated result images with contours and a metrics table including average RGB values.
- Semi-transparent background for annotations to enhance readability.
- Additional filtering based on average RGB values.
- Dynamic table column width adjustment.
- Option to remove "mm²" from the area column.
- Comprehensive logging to file and console.

Author:
-------
Rasmus Jensen
raje at ecos au dk

Date:
-----
October 29, 2024

Version:
--------
1.0.0

"""

import os
import sys
import time
import logging
import configparser
from dataclasses import dataclass
from typing import List, Tuple, Dict, Optional, Any

import cv2
import numpy as np
import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from PIL import Image, ImageDraw, ImageFont
import re
import threading
import ast

# ============================ Configuration and Logging Setup ============================

# Initialize Config
current_dir = os.path.dirname(os.path.abspath(__file__))
config_path = os.path.join(current_dir, '..', 'config', 'config.ini')  # Adjusted path based on directory structure
config = configparser.ConfigParser()

# Read configuration from config.ini
if os.path.exists(config_path):
    config.read(config_path)
    print(f"Configuration loaded from {config_path}.")
    logging.info(f"Configuration loaded from {config_path}.")
else:
    print(f"Configuration file not found at {config_path}. Using default settings.")
    logging.warning(f"Configuration file not found at {config_path}. Using default settings.")
    # Optionally, create a default config.ini here or handle as needed

# Set up logging based on configuration
log_level_str = config.get("DEFAULT", "log_level", fallback="DEBUG")
log_level = getattr(logging, log_level_str.upper(), logging.DEBUG)

print(f"Logging level set to: {log_level_str}")

# Configure logging to file and console
logging.basicConfig(level="DEBUG",
                    format='%(asctime)s - %(levelname)s - %(message)s',
                    handlers=[
                        logging.FileHandler(os.path.join(current_dir, 'processing.log')),
                        logging.StreamHandler(sys.stdout)
                    ])
logger = logging.getLogger(__name__)
logger.debug("Logging is configured.")

# ============================ Data Classes ============================

@dataclass
class Config:
    """
    Configuration dataclass to hold all configuration parameters.

    Attributes:
        image_directory (str): Directory containing leaf images.
        area_threshold (float): Minimum leaf area in mm² to consider.
        crop_left_pct (float): Percentage to crop from the left.
        crop_right_pct (float): Percentage to crop from the right.
        crop_top_pct (float): Percentage to crop from the top.
        crop_bottom_pct (float): Percentage to crop from the bottom.
        filename (str): Specific filename to process.
        img_debug (bool): Flag to save intermediate images.
        adaptive_threshold (bool): Enable adaptive thresholding in RGB space.
        adaptive_window_size (int): Window size for adaptive thresholding.
        adaptive_C (int): Constant subtracted from mean in adaptive thresholding.
        color_threshold (int): Threshold to define near-white pixels.
        kernel_size (Tuple[int, int]): Kernel size for morphological operations.
        filter_rgb (bool): Flag to enable additional RGB-based filtering.
        r_threshold (int): Threshold for Red channel.
        g_threshold (int): Threshold for Green channel.
        b_threshold (int): Threshold for Blue channel.
    """
    image_directory: str
    area_threshold: float
    crop_left_pct: float
    crop_right_pct: float
    crop_top_pct: float
    crop_bottom_pct: float
    filename: str
    img_debug: bool
    adaptive_threshold: bool
    adaptive_window_size: int
    adaptive_C: int
    color_threshold: int
    kernel_size: Tuple[int, int]
    filter_rgb: bool
    r_threshold: int
    g_threshold: int
    b_threshold: int

# ============================ Configuration Reader ============================

def read_config(config_file: str = 'config.ini') -> Config:
    """
    Read the configuration from 'config.ini' and return a Config object.

    Args:
        config_file (str): Name of the configuration file.

    Returns:
        Config: Configuration object with all parameters.
    """
    # Determine the path to the config.ini file relative to this script
    script_dir = os.path.dirname(os.path.abspath(__file__))
    config_path = os.path.join(script_dir, '..', 'config', config_file)

    config_parser = configparser.ConfigParser()
    config_parser.read(config_path)
    config = config_parser['DEFAULT']

    # Parse kernel_size safely
    kernel_size_str = config.get('kernel_size', fallback='(5, 5)')
    try:
        kernel_size = ast.literal_eval(kernel_size_str)
        if not (isinstance(kernel_size, tuple) and len(kernel_size) == 2 and all(isinstance(x, int) for x in kernel_size)):
            raise ValueError
    except:
        logger.error(f"Invalid kernel_size format in config.ini. Using default (5, 5).")
        kernel_size = (5, 5)

    return Config(
        image_directory=config.get('image_directory', ''),
        area_threshold=config.getfloat('area_threshold', fallback=2.0),
        crop_left_pct=config.getfloat('crop_left', fallback=10.0) / 100.0,
        crop_right_pct=config.getfloat('crop_right', fallback=2.0) / 100.0,
        crop_top_pct=config.getfloat('crop_top', fallback=2.0) / 100.0,
        crop_bottom_pct=config.getfloat('crop_bottom', fallback=2.0) / 100.0,
        filename=config.get('filename', fallback=''),
        img_debug=config.getboolean('img_debug', fallback=False),
        adaptive_threshold=config.getboolean('adaptive_threshold', fallback=True),
        adaptive_window_size=config.getint('adaptive_window_size', fallback=15),
        adaptive_C=config.getint('adaptive_C', fallback=2),
        color_threshold=config.getint('color_threshold', fallback=200),
        kernel_size=kernel_size,
        filter_rgb=config.getboolean('filter_rgb', fallback=True),
        r_threshold=config.getint('r_threshold', fallback=200),
        g_threshold=config.getint('g_threshold', fallback=200),
        b_threshold=config.getint('b_threshold', fallback=200)
    )

# ============================ Directory Setup ============================

def create_directories(config: Config) -> Tuple[str, str]:
    """
    Create necessary directories for intermediate and result images.

    Args:
        config (Config): Configuration object.

    Returns:
        Tuple[str, str]: Paths to the intermediate and result image directories.
    """
    result_folder = os.path.join(config.image_directory, 'result_img')
    intermediate_folder = os.path.join(result_folder, 'intermediate_img')

    if not os.path.exists(result_folder):
        os.makedirs(result_folder)
    logger.debug(f"Created result folder at: {result_folder}")

    if config.img_debug:
        if not os.path.exists(intermediate_folder):
            os.makedirs(intermediate_folder)
        logger.debug(f"Created intermediate folder at: {intermediate_folder}")

    return intermediate_folder, result_folder

# ============================ Image File Retrieval ============================

def get_image_files(config: Config) -> List[str]:
    """
    Get the list of image files to process.

    Args:
        config (Config): Configuration object.

    Returns:
        List[str]: List of image filenames.
    """
    if config.filename:
        logger.debug(f"Processing single file: {config.filename}")
        return [config.filename]
    else:
        valid_extensions = ('.jpg', '.jpeg', '.png', '.tif', '.tiff', '.bmp')
        image_files = [f for f in os.listdir(config.image_directory)
                       if f.lower().endswith(valid_extensions)]
        logger.debug(f"Found {len(image_files)} image files.")
        return image_files

# ============================ Image Processing Functions ============================

def process_images_main(stop_event: threading.Event) -> None:
    """
    Main function to process images based on configuration settings.

    Args:
        stop_event (threading.Event): Event to signal stopping of processing.
    """
    try:
        # Read configuration
        config = read_config()

        # Create necessary directories
        intermediate_folder, result_folder = create_directories(config)

        # Get list of image files to process
        image_files = get_image_files(config)

        # Define paths for results
        csv_file_path = os.path.join(config.image_directory, 'leaf_analysis_results.csv')

        # Initialize list to hold results
        results_list = []

        # Iterate through each image file
        for image_file in image_files:
            if stop_event.is_set():
                logger.info("Processing was stopped before processing all images.")
                break

            image_path = os.path.join(config.image_directory, image_file)
            result = process_image(image_path, config, intermediate_folder, result_folder, stop_event)

            if result:
                results_list.append(result)

        # Save results if processing wasn't interrupted
        if results_list and not (stop_event.is_set()):
            save_results(results_list, csv_file_path)
        elif stop_event.is_set():
            logger.info("Processing was interrupted. Partial results may not be saved.")

    except Exception as e:
        logger.error(f"An unexpected error occurred in process_images_main: {e}")
    finally:
        logger.info("Image processing completed.")

def process_image(image_path: str, config: Config, intermediate_folder: str, result_folder: str,
                  stop_event: threading.Event) -> Optional[Dict[str, Any]]:
    filename = os.path.basename(image_path)
    cropped_cv = None  # Initialize variable

    try:
        logger.debug(f"Loading image: {image_path}")
        # Load the image using PIL first
        pil_image = Image.open(image_path)
        width, height = pil_image.size
        dpi = pil_image.info.get('dpi', (600, 600))
        x_dpi = dpi[0] if dpi else 600
        pixels_per_mm = x_dpi / 25.4  # Convert DPI to pixels per millimeter

        logger.debug(f"Image size: {width}x{height}, DPI: {dpi}, Pixels/mm: {pixels_per_mm:.2f}")

        # Convert to OpenCV format (BGR)
        image_cv = cv2.cvtColor(np.array(pil_image), cv2.COLOR_RGB2BGR)
    
        # Apply image preprocessing steps
        processed_image, cropped_cv, top_offset, left_offset = preprocess_image(
            image_cv, config, intermediate_folder, filename, pixels_per_mm, stop_event
        )
        if processed_image is None or cropped_cv is None:
            logger.error(f"Preprocessing failed for {filename}")
            return

    except Exception as e:
        logger.error(f"Error processing {filename}: {e}", exc_info=True)
        return

    if stop_event.is_set():
        logger.info("Processing was stopped before starting.")
        return None

    start_time = time.time()

    # Proceed with further processing using 'cropped_cv'

    try:
        logger.info(f"Processing {filename}")

        # Open the image and get its properties
        pil_image = Image.open(image_path)
        width, height = pil_image.size
        dpi = pil_image.info.get('dpi', (600, 600))
        x_dpi = dpi[0] if dpi else 600
        pixels_per_mm = x_dpi / 25.4  # Convert DPI to pixels per millimeter

        logger.debug(f"Image size: {width}x{height} pixels, DPI: {dpi}, Pixels/mm: {pixels_per_mm:.2f}")

        # Convert to OpenCV format (BGR)
        image_cv = cv2.cvtColor(np.array(pil_image), cv2.COLOR_RGB2BGR)

        # Apply image preprocessing steps
        processed_image, cropped_cv, top_offset, left_offset = preprocess_image(
            image_cv, config, intermediate_folder, filename, pixels_per_mm, stop_event
        )

        if stop_event.is_set():
            logger.info(f"Processing was stopped after preprocessing {filename}.")
            return None

        # Find and process contours, including RGB stats
        total_count, total_area_mm2, table, leaf_rgb_stats = find_and_process_contours(
            processed_image, cropped_cv, pixels_per_mm, config, result_folder,
            filename, top_offset, left_offset, stop_event
        )

        if stop_event.is_set():
            logger.info(f"Processing was stopped after contour processing {filename}.")
            return None

        processing_time = time.time() - start_time
        logger.info(f"Finished processing {filename}. "
                    f"Leaf Count: {total_count}, Total Leaf Area: {total_area_mm2:.2f} mm², "
                    f"Processing Time: {processing_time:.2f} s")

        # Perform regex searches once and reuse
        pct_match = re.search(r"_(HF|CF)_", filename)
        leaf_stem_match = re.search(r"_(L|S)_", filename)

        # Print table content to console
        if table:
            print(f"\n--- Leaf Analysis Table for {filename} ---")
            print(f"{'ID':<10}{'Area':<15}{'Average RGB':<15}")
            for row in table:
                print(f"{row[0]:<10}{row[1]:<15}{row[2]:<15}")
            print(f"---------------------------------------------\n")
        else:
            print(f"No leaves detected in {filename}.")

        return {
            'Filename': filename,
            'PCT': pct_match.group(1) if pct_match else None,
            'Leaf_stem': leaf_stem_match.group(1) if leaf_stem_match else None,
            'Leaf Count': total_count,
            'Total Leaf Area (mm²)': total_area_mm2,
            'Pixels_per_mm': pixels_per_mm,
            'Processing Time (s)': round(processing_time, 2),
            'Leaf RGB Stats': leaf_rgb_stats  # Include RGB stats
        }

    except Exception as e:
        logger.error(f"Error processing {filename}: {e}")
        return
    # Ensure 'cropped_cv' is assigned before further processing
    if cropped_cv is not None:
        # Proceed with processing using 'cropped_cv'
        pass
    else:
        logger.error("cropped_cv is not assigned. Cannot proceed with processing.")
        return

def preprocess_image(image_cv, config, intermediate_folder, filename, pixels_per_mm, stop_event):
    # Initialize variables
    cropped_cv = None
    cleaned_image = None  # Assuming this is also used

    try:
        # Step 1: Crop the image
        # Implement your cropping logic here
        cropped_image = image_cv[...]
        cropped_cv = cropped_image.copy()
        logger.debug("Image cropped successfully.")

        # Proceed with processing to get 'cleaned_image'
        # For example:
        cleaned_image = cv2.medianBlur(cropped_cv, 5)  # Example cleaning operation using median blur
        logger.debug("Image cleaned successfully.")

    except Exception as e:
        logger.error(f"Error in preprocess_image: {e}", exc_info=True)
        return None, None, None, None

    if cropped_cv is None or cleaned_image is None:
        logger.error("Required images are not available for size comparison.")
        return None, None, None, None

    # Now safe to compare sizes
    if cleaned_image.shape[:2] != cropped_cv.shape[:2]:
        # Handle size mismatch
        logger.error("Size mismatch between cleaned_image and cropped_cv.")
        return None, None, None, None

    # Continue with the rest of the function
    # ...

    # Return the required values
    processed_image = cleaned_image  # Assuming processed_image is the cleaned_image
    top_offset = 0  # Define top_offset appropriately
    left_offset = 0  # Define left_offset appropriately
    return processed_image, cropped_cv, top_offset, left_offset

def remove_small_objects(closed_image: np.ndarray, area_threshold_pixels: float) -> np.ndarray:
    """
    Remove small objects from binary image based on area threshold.

    Args:
        closed_image (np.ndarray): Binary image after morphological closing.
        area_threshold_pixels (float): Area threshold in pixels.

    Returns:
        np.ndarray: Cleaned binary image with small objects removed.
    """
    # Find connected components
    num_labels, labels_im = cv2.connectedComponents(closed_image)

    # Count the size of each component
    sizes = np.bincount(labels_im.ravel())

    # Create a mask for components that meet the area threshold
    mask_sizes = sizes >= area_threshold_pixels
    mask_sizes[0] = 0  # Ensure background is not included

    # Apply the mask to the labels image
    output_image = mask_sizes[labels_im]

    # Convert the boolean mask to uint8
    output_image = (output_image * 255).astype(np.uint8)

    logger.debug(f"Removed small objects below {area_threshold_pixels} pixels.")
    return output_image


def save_image(image: np.ndarray, folder: str, filename: str, step_name: str,
               img_debug: bool, background_image: Optional[np.ndarray] = None) -> None:
    if img_debug:
        try:
            # Check if the processed image is valid
            if image is None or image.size == 0:
                logger.error(f"Processed image is empty for {filename} at step {step_name}")
                return

            # Check if the background image is valid (if provided)
            if background_image is not None and (background_image.size == 0):
                logger.error(f"Background image is empty for {filename} at step {step_name}")
                return

            if background_image is not None:
                # Resize the background image to match the processed image dimensions
                new_width = int(image.shape[1])
                new_height = int(image.shape[0])
                background_image = cv2.resize(background_image, (new_width, new_height))
                logger.debug(f"Background image resized to: {background_image.shape}")

                # Convert background image to BGR if needed
                if len(background_image.shape) == 2:
                    background_image = cv2.cvtColor(background_image, cv2.COLOR_GRAY2BGR)
                elif background_image.shape[2] == 4:
                    background_image = cv2.cvtColor(background_image, cv2.COLOR_BGRA2BGR)

                # Convert processed image to BGR if needed
                if len(image.shape) == 2:
                    image_colored = cv2.cvtColor(image, cv2.COLOR_GRAY2BGR)
                elif image.shape[2] == 4:
                    image_colored = cv2.cvtColor(image, cv2.COLOR_BGRA2BGR)
                else:
                    image_colored = image

                # Log shapes and data types
                logger.debug(f"image_colored shape: {image_colored.shape}, dtype: {image_colored.dtype}")
                logger.debug(f"background_image shape: {background_image.shape}, dtype: {background_image.dtype}")

                # Ensure both images have the same dimensions
                if image_colored.shape != background_image.shape:
                    logger.error(f"Dimension mismatch: image_colored shape {image_colored.shape}, background_image shape {background_image.shape}")
                    return

                # Ensure both images have the same data type
                if image_colored.dtype != background_image.dtype:
                    logger.debug(f"Data type mismatch: Converting both images to 'uint8'")
                    image_colored = image_colored.astype('uint8')
                    background_image = background_image.astype('uint8')

                # Blend the images
                alpha = 0.5  # Transparency factor for the processed image
                faded_image = cv2.addWeighted(image_colored, alpha, background_image, 1 - alpha, 0)

                # Save the faded image with the background image
                image_name = f"{os.path.splitext(filename)[0]}_{step_name}.png"
                image_path = os.path.join(folder, image_name)
                cv2.imwrite(image_path, faded_image)
                logger.debug(f"Saved intermediate image with background: {image_path}")
            else:
                # No background image provided; save the processed image as is
                image_name = f"{os.path.splitext(filename)[0]}_{step_name}.png"
                image_path = os.path.join(folder, image_name)

                # Convert processed image to BGR if needed
                if len(image.shape) == 2:
                    image_colored = cv2.cvtColor(image, cv2.COLOR_GRAY2BGR)
                elif image.shape[2] == 4:
                    image_colored = cv2.cvtColor(image, cv2.COLOR_BGRA2BGR)
                else:
                    image_colored = image

                cv2.imwrite(image_path, image_colored)
                logger.debug(f"Saved intermediate image: {image_path}")
        except Exception as e:
            logger.error(f"Failed to save image {filename} at step {step_name}: {e}", exc_info=True)

def save_intermediate_image(image: np.ndarray, step_name: str, folder: str, filename: str, background_image: Optional[np.ndarray]=None):
    try:
        os.makedirs(folder, exist_ok=True)
        image_name = f"{os.path.splitext(filename)[0]}_{step_name}.png"
        image_path = os.path.join(folder, image_name)
        if background_image is not None:
            # Blend with background image
            alpha = 0.5
            blended_image = cv2.addWeighted(image, alpha, background_image, 1 - alpha, 0)
            cv2.imwrite(image_path, blended_image)
        else:
            cv2.imwrite(image_path, image)
        logger.debug(f"Saved intermediate image: {image_path}")
    except Exception as e:
        logger.error(f"Failed to save image {filename} at step {step_name}: {e}", exc_info=True)

def find_and_process_contours(processed_image: np.ndarray, cropped_cv: np.ndarray, pixels_per_mm: float,
                              config: Config, result_folder: str, filename: str,
                              top_offset: int, left_offset: int, stop_event: threading.Event) -> Tuple[int, float, List[Tuple[str, str, str]], List[Dict[str, Any]]]:
    if stop_event.is_set():
        logger.info("Processing was stopped before contour processing.")
        return 0, 0.0, [], []

    # Step 1: Convert the processed image to grayscale if it is not already
    if len(processed_image.shape) == 3:
        logger.debug("Converting processed_image to grayscale.")
        processed_gray = cv2.cvtColor(processed_image, cv2.COLOR_BGR2GRAY)
    else:
        processed_gray = processed_image

    # Step 2: Apply thresholding to get a binary image
    logger.debug("Applying thresholding to get a binary image.")
    _, binary_image = cv2.threshold(processed_gray, 0, 255, cv2.THRESH_BINARY + cv2.THRESH_OTSU)

    # Save the binary image for inspection
    binary_image_path = os.path.join(result_folder, f"{os.path.splitext(filename)[0]}_binary.png")
    cv2.imwrite(binary_image_path, binary_image)
    logger.debug(f"Saved binary image: {binary_image_path}")

    # Step 3: Find contours on the binary image
    logger.debug("Finding contours.")
    contours, _ = cv2.findContours(binary_image, cv2.RETR_EXTERNAL, cv2.CHAIN_APPROX_SIMPLE)
    logger.debug(f"Found {len(contours)} contours.")

    if not contours:
        logger.warning("No contours found.")
        return 0, 0.0, [], []

    # Proceed with contour processing
    total_count = 0
    total_area = 0.0
    pixels_per_mm2 = pixels_per_mm ** 2
    output_image = cropped_cv.copy()
    annotated_images = []
    measurement_data = []

    try:
        # Process each contour
        for contour in contours:
            # Calculate area in pixels
            area_pixels = cv2.contourArea(contour)
            # Convert area to square millimeters
            area_mm2 = area_pixels / pixels_per_mm2

            # Filter out small areas if necessary
            if area_mm2 < config.min_leaf_area_mm2:
                logger.debug(f"Contour area {area_mm2:.2f} mm² is smaller than the minimum threshold.")
                continue

            total_count += 1
            total_area += area_mm2

            # Draw contour on the output image
            cv2.drawContours(output_image, [contour], -1, (0, 255, 0), 2)

            # Additional processing (e.g., compute average RGB values)
            # ...

        # Save the output image with contours
        output_image_path = os.path.join(result_folder, f"{os.path.splitext(filename)[0]}_contours.png")
        cv2.imwrite(output_image_path, output_image)
        logger.debug(f"Saved output image with contours: {output_image_path}")

    except Exception as e:
        logger.error(f"Error during contour processing: {e}", exc_info=True)
        return 0, 0.0, [], []

    return total_count, total_area, annotated_images, measurement_data


def compute_average_rgb(image_cv: np.ndarray, contour: np.ndarray) -> Tuple[int, int, int]:
    """
    Compute the average RGB values within a given contour.

    Args:
        image_cv (np.ndarray): Cropped OpenCV image in BGR format.
        contour (np.ndarray): Contour of the leaf.

    Returns:
        Tuple[int, int, int]: Average (R, G, B) values.
    """
    # Create a mask for the contour
    mask = np.zeros(image_cv.shape[:2], dtype=np.uint8)
    cv2.drawContours(mask, [contour], -1, 255, -1)  # Filled contour

    # Convert image to RGB
    image_rgb = cv2.cvtColor(image_cv, cv2.COLOR_BGR2RGB)

    # Calculate mean R, G, B values within the mask
    mean_r = cv2.mean(image_rgb[:, :, 0], mask=mask)[0]
    mean_g = cv2.mean(image_rgb[:, :, 1], mask=mask)[0]
    mean_b = cv2.mean(image_rgb[:, :, 2], mask=mask)[0]

    return int(mean_r), int(mean_g), int(mean_b)


def extract_rgb_within_contour(image_cv: np.ndarray, contour: np.ndarray) -> np.ndarray:
    """
    Extract RGB values within a given contour.

    Args:
        image_cv (np.ndarray): Cropped OpenCV image in BGR format.
        contour (np.ndarray): Contour of the leaf.

    Returns:
        np.ndarray: RGB image masked by the contour.
    """
    # Create a mask for the contour
    mask = np.zeros(image_cv.shape[:2], dtype=np.uint8)
    cv2.drawContours(mask, [contour], -1, 255, -1)  # Filled contour

    # Convert image to RGB
    image_rgb = cv2.cvtColor(image_cv, cv2.COLOR_BGR2RGB)

    # Apply mask
    masked_rgb = cv2.bitwise_and(image_rgb, image_rgb, mask=mask)

    return masked_rgb


def create_annotation_table(annotations: List[Dict[str, Any]], areas_mm2: List[float],
                            leaf_rgb_stats: List[Dict[str, Any]]) -> List[Tuple[str, str, str]]:
    """
    Create a table of annotations with contour numbers, areas, and average RGB values.

    Args:
        annotations (List[Dict[str, Any]]): List of annotation dictionaries.
        areas_mm2 (List[float]): List of leaf areas in mm².
        leaf_rgb_stats (List[Dict[str, Any]]): List of RGB statistics dictionaries.

    Returns:
        List[Tuple[str, str, str]]: List of tuples representing table rows.
    """
    table = []
    for ann, area, rgb_stat in zip(annotations, areas_mm2, leaf_rgb_stats):
        average_rgb = f"{int(rgb_stat['Mean R'])}/{int(rgb_stat['Mean G'])}/{int(rgb_stat['Mean B'])}"
        table.append((ann['text'], f"{area:.2f}", average_rgb))
    logger.debug("Created annotation table with average RGB values.")
    return table


def add_table_to_image(image: np.ndarray, table: List[Tuple[str, str, str]], table_position: Tuple[int, int] = (50, 50)) -> np.ndarray:
    """
    Add a table to the image at the specified position with enhanced aesthetics.

    Args:
        image (np.ndarray): Image to draw the table on.
        table (List[Tuple[str, str, str]]): List of tuples representing table rows.
        table_position (Tuple[int, int], optional): (x, y) position to place the table.

    Returns:
        np.ndarray: Image with the table added.
    """
    font = cv2.FONT_HERSHEY_SIMPLEX
    base_font_scale = 0.6  # Base font scale for visibility
    base_font_thickness = 2
    base_line_height = 30
    base_margin = 15

    # Determine table size based on image width to ensure at least 10% width
    image_height, image_width = image.shape[:2]
    min_table_width = int(0.1 * image_width)
    current_table_width = 450  # Increased width for better visibility
    table_width = max(current_table_width, min_table_width)

    # Calculate the number of rows and columns
    num_rows = len(table) + 1  # Including header
    num_cols = 3  # ID, Area, Average RGB
    table_height = num_rows * base_line_height + 2 * base_margin

    x_start, y_start = table_position

    # Dynamic Font Scaling based on image width
    base_width = 1920  # Reference width for scaling
    scale_factor = image_width / base_width
    font_scale = base_font_scale * scale_factor
    font_thickness = max(1, int(base_font_thickness * scale_factor))
    line_height = int(base_line_height * scale_factor)
    margin = int(base_margin * scale_factor)

    # Adjust table position if it exceeds image boundaries
    if x_start + table_width > image_width:
        x_start = image_width - table_width - 20  # 20 pixels padding from the edge
    if y_start + table_height > image_height:
        y_start = image_height - table_height - 20  # 20 pixels padding from the edge

    logger.debug(f"Table position: ({x_start}, {y_start}), Width: {table_width}px, Height: {table_height}px")

    # Draw semi-transparent table background
    overlay = image.copy()
    cv2.rectangle(overlay, (x_start, y_start),
                  (x_start + table_width, y_start + table_height),
                  (255, 255, 255), -1)  # White background
    alpha = 0.6  # Transparency factor
    cv2.addWeighted(overlay, alpha, image, 1 - alpha, 0, image)

    # Draw table border
    cv2.rectangle(image, (x_start, y_start),
                  (x_start + table_width, y_start + table_height),
                  (0, 0, 0), 2)  # Black border

    # Define column headers
    headers = ["ID", "Area", "Average RGB"]

    # Initialize lists to store maximum text widths for each column
    max_text_widths = [0, 0, 0]

    # Calculate maximum text width for each column based on headers and data
    for i in range(num_cols):
        header_text = headers[i]
        (header_width, _), _ = cv2.getTextSize(header_text, font, font_scale, font_thickness)
        max_text_widths[i] = header_width + 20  # Adding padding

    for row in table:
        for i in range(num_cols):
            (text_width, _), _ = cv2.getTextSize(row[i], font, font_scale, font_thickness)
            if text_width + 20 > max_text_widths[i]:
                max_text_widths[i] = text_width + 20  # Update if current text is wider

    # Calculate total required table width
    required_table_width = sum(max_text_widths)
    if required_table_width > table_width:
        table_width = required_table_width
        # Redraw semi-transparent background and border with updated table width
        overlay = image.copy()
        cv2.rectangle(overlay, (x_start, y_start),
                      (x_start + table_width, y_start + table_height),
                      (255, 255, 255), -1)  # White background
        alpha = 0.6  # Transparency factor
        cv2.addWeighted(overlay, alpha, image, 1 - alpha, 0, image)
        cv2.rectangle(image, (x_start, y_start),
                      (x_start + table_width, y_start + table_height),
                      (0, 0, 0), 2)  # Black border

    # Recalculate column widths if necessary
    col_widths = max_text_widths
    total_col_width = sum(col_widths)

    # Adjust column widths proportionally if they exceed the table width
    if total_col_width > table_width:
        scale = table_width / total_col_width
        col_widths = [int(width * scale) for width in col_widths]

    # Add header
    for i, text in enumerate(headers):
        x_offset = sum(col_widths[:i]) + 10  # Adding padding
        y_offset = y_start + margin + line_height
        cv2.putText(image, text, (x_start + x_offset, y_start + margin + line_height),
                    font, font_scale, (0, 0, 0), font_thickness, cv2.LINE_AA)

    # Add rows with alternating background colors for better readability
    for idx, row in table:
        y = y_start + margin + (idx + 2) * line_height  # +2 to account for header
        # Alternate row color
        if idx % 2 == 0:
            row_color = (240, 240, 240)  # Light gray
            cv2.rectangle(image, (x_start, y - line_height + 5),
                          (x_start + table_width, y + 5),
                          row_color, -1)
        for col_idx in range(num_cols):
            x_offset = sum(col_widths[:col_idx]) + 10  # Adding padding
            display_text = row[col_idx]
            cv2.putText(image, display_text, (x_start + x_offset, y),
                        font, font_scale, (0, 0, 0), font_thickness, cv2.LINE_AA)

    # Draw vertical lines for column separators
    current_x = x_start
    for width in col_widths[:-1]:  # No line after the last column
        current_x += width
        cv2.line(image, (current_x, y_start),
                 (current_x, y_start + table_height),
                 (0, 0, 0), 2)

    logger.debug("Added table to image with enhanced aesthetics.")
    return image


def save_results(results_list: List[Dict[str, Any]], csv_file_path: str) -> None:
    """
    Save the analysis results, including mean RGB values, to a CSV file and an Excel file.
    
    Args:
        results_list (List[Dict[str, Any]]): List of dictionaries with analysis results.
        csv_file_path (str): Path to the CSV file.
    """
    # Prepare data for CSV
    rows = []
    for result in results_list:
        base_info = {
            'Filename': result.get('Filename'),
            'PCT': result.get('PCT'),
            'Leaf_stem': result.get('Leaf_stem'),
            'Leaf Count': result.get('Leaf Count'),
            'Total Leaf Area (mm²)': result.get('Total Leaf Area (mm²)'),
            'Pixels_per_mm': result.get('Pixels_per_mm'),
            'Processing Time (s)': result.get('Processing Time (s)')
        }
        # Add RGB stats
        for leaf_stat in result.get('Leaf RGB Stats', []):
            row = base_info.copy()
            row['Leaf ID'] = leaf_stat['Leaf ID']
            row['Leaf Area'] = leaf_stat['Leaf Area']
            row['Average R'] = leaf_stat['Mean R']
            row['Average G'] = leaf_stat['Mean G']
            row['Average B'] = leaf_stat['Mean B']
            row['Average RGB'] = f"{leaf_stat['Mean R']}/{leaf_stat['Mean G']}/{leaf_stat['Mean B']}"
            rows.append(row)

    df_csv = pd.DataFrame(rows)

    if os.path.exists(csv_file_path):
        try:
            df_existing = pd.read_csv(csv_file_path)
            df_combined = pd.concat([df_existing, df_csv], ignore_index=True)
        except pd.errors.EmptyDataError:
            df_combined = df_csv
    else:
        df_combined = df_csv

    df_combined.to_csv(csv_file_path, index=False)
    logger.info(f"Results saved to CSV: {csv_file_path}")

     # Define the Excel file path
    excel_file_path = os.path.splitext(csv_file_path)[0] + '.xlsx'

    # Prepare data for "Raw_output" sheet
    raw_rows = []
    for result in results_list:
        filename = result.get('Filename')
        pct = result.get('PCT')
        leaf_stem = result.get('Leaf_stem')
        # Iterate through each leaf's stats
        for leaf_stat in result.get('Leaf RGB Stats', []):
            row = {
                'Filename': filename,
                'PCT': pct,
                'Leaf_stem': leaf_stem,
                'Leaf ID': leaf_stat['Leaf ID'],
                'Leaf Area': leaf_stat['Leaf Area'],  # Use exact Leaf Area from stats
                'Average R': leaf_stat['Mean R'],
                'Average G': leaf_stat['Mean G'],
                'Average B': leaf_stat['Mean B']
            }
            raw_rows.append(row)
    df_raw = pd.DataFrame(raw_rows)

    # Write "Raw_output" and "Summary" sheets to Excel
    with pd.ExcelWriter(excel_file_path, engine='openpyxl') as writer:
        # Write Raw_output sheet
        df_raw.to_excel(writer, sheet_name='Raw_output', index=False)

        # Create Summary sheet with headers
        summary_headers = ['Filename', 'PCT', 'Leaf_stem', 'Total Leaf Area', 'Average R', 'Average G', 'Average B']
        df_summary = pd.DataFrame(columns=summary_headers)
        df_summary.to_excel(writer, sheet_name='Summary', index=False)

        # Access the workbook and sheets
        workbook = writer.book
        raw_sheet = writer.sheets['Raw_output']
        summary_sheet = writer.sheets['Summary']

        # Get unique combinations of Filename, PCT, Leaf_stem
        unique_images = df_raw[['Filename', 'PCT', 'Leaf_stem']].drop_duplicates().reset_index(drop=True)

        for idx, row in unique_images.iterrows():
            excel_row = idx + 2  # Excel rows start at 1, header is row 1
            filename = row['Filename']
            pct = row['PCT']
            leaf_stem = row['Leaf_stem']

            # Write Filename, PCT, Leaf_stem to Summary sheet
            summary_sheet.cell(row=excel_row, column=1, value=filename)
            summary_sheet.cell(row=excel_row, column=2, value=pct)
            summary_sheet.cell(row=excel_row, column=3, value=leaf_stem)

            # Define the range for SUMIF and SUMPRODUCT
            # Raw_output sheet columns:
            # A: Filename, B: PCT, C: Leaf_stem, D: Leaf ID, E: Leaf Area, F: Average R, G: Average G, H: Average B
            total_leaf_area_formula = f"=SUMIF(Raw_output!A:A, Summary!A{excel_row}, Raw_output!E:E)"
            average_r_formula = f"=SUMPRODUCT(--(Raw_output!A2:A1000=Summary!A{excel_row}), Raw_output!E2:E1000, Raw_output!F2:F1000) / Summary!D{excel_row}"
            average_g_formula = f"=SUMPRODUCT(--(Raw_output!A2:A1000=Summary!A{excel_row}), Raw_output!E2:E1000, Raw_output!G2:G1000) / Summary!D{excel_row}"
            average_b_formula = f"=SUMPRODUCT(--(Raw_output!A2:A1000=Summary!A{excel_row}), Raw_output!E2:E1000, Raw_output!H2:H1000) / Summary!D{excel_row}"

            # Write formulas to Summary sheet
            summary_sheet.cell(row=excel_row, column=4, value=total_leaf_area_formula)
            summary_sheet.cell(row=excel_row, column=5, value=average_r_formula)
            summary_sheet.cell(row=excel_row, column=6, value=average_g_formula)
            summary_sheet.cell(row=excel_row, column=7, value=average_b_formula)

        # Adjust column widths for better readability
        for sheet_name in ['Raw_output', 'Summary']:
            sheet = writer.sheets[sheet_name]
            for column_cells in sheet.columns:
                length = max(len(str(cell.value)) for cell in column_cells)
                adjusted_width = (length + 2)
                sheet.column_dimensions[get_column_letter(column_cells[0].column)].width = adjusted_width

    logger.info(f"Results saved to Excel: {excel_file_path}")


def main(stop_event: Optional[threading.Event] = None) -> None:
    """
    Main function to orchestrate the leaf classification process.

    This function performs the following steps:
    1. Reads configuration settings from 'config.ini'.
    2. Creates necessary directories for storing intermediate and result images.
    3. Identifies image files to process based on configuration.
    4. Iterates through each image file, processing them individually.
    5. Accumulates results and saves them to CSV and Excel files.
    6. Handles graceful termination if a stop event is signaled.

    Args:
        stop_event (Optional[threading.Event], optional):
            Event to signal stopping of processing. If set, the function
            will terminate processing after the current image.
    """
    try:
        # Step 1: Read configuration settings
        config = read_config()
        logger.debug("Configuration settings loaded successfully.")

        # Step 2: Create directories for intermediate and result images
        intermediate_folder, result_folder = create_directories(config)
        logger.debug(f"Intermediate images will be saved to: {intermediate_folder}")
        logger.debug(f"Processed result images will be saved to: {result_folder}")

        # Define the path for the CSV results file
        csv_file_path = os.path.join(config.image_directory, 'leaf_analysis_results.csv')
        logger.debug(f"Results will be saved to CSV file at: {csv_file_path}")

        # Step 3: Get list of image files to process
        image_files = get_image_files(config)
        logger.info(f"Found {len(image_files)} image(s) to process.")

        # Initialize a list to store results from each image
        results_list = []

        # Step 4: Iterate through each image file and process
        for image_file in image_files:
            # Check if a stop signal has been received
            if stop_event and stop_event.is_set():
                logger.info("Processing was stopped before processing all images.")
                break

            # Construct the full path to the image file
            image_path = os.path.join(config.image_directory, image_file)
            logger.debug(f"Starting processing for image: {image_path}")

            # Process the individual image
            result = process_image(image_path, config, intermediate_folder, result_folder, stop_event)

            # If processing was successful, append the result to the results list
            if result:
                results_list.append(result)
                logger.debug(f"Result appended for image: {image_file}")
            else:
                logger.warning(f"No result returned for image: {image_file}")

        # Step 5: Save accumulated results if processing wasn't stopped
        if results_list and not (stop_event and stop_event.is_set()):
            save_results(results_list, csv_file_path)
            logger.info("All results have been saved successfully.")
        elif stop_event and stop_event.is_set():
            logger.info("Processing was interrupted. Partial results may not be saved.")

    except Exception as e:
        # Catch any unexpected exceptions and log them
        logger.error(f"An unexpected error occurred in main(): {e}", exc_info=True)
    finally:
        # Final log entry to indicate that the processing has concluded
        logger.info("Leaf classification processing has concluded.")


if __name__ == "__main__":
    main()
